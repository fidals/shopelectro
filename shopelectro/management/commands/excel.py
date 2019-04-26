"""
Generate Excel price-list.

Use this excel editor lib: https://openpyxl.readthedocs.io/en/stable/
"""
import datetime
import os
from collections import namedtuple

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl.styles import borders, colors, Font

from shopelectro.models import Product, Category


class Command(BaseCommand):
    TEMPLATE = 'templates/ecommerce/template.xlsx'
    NAME = 'pricelist.xlsx'
    SHEET_TITLE = 'Прайс Shopelectro'
    CATEGORY_FILL = openpyxl.styles.PatternFill(
        start_color='F4FEFD',
        end_color='F4FEFD',
        fill_type='solid'
    )
    BUY_FILL = openpyxl.styles.PatternFill(
        start_color='FEFEF0',
        end_color='FEFEF0',
        fill_type='solid'
    )
    THIN_BORDER = borders.Border(
        top=borders.Side(style='thin'),
        right=borders.Side(style='thin'),
        bottom=borders.Side(style='thin'),
        left=borders.Side(style='thin')
    )
    CURRENT_ROW = '9'  # Start of catalog section in file.
    cell = namedtuple('cell', ['row', 'col'])
    BAD_STYLED_CELLS = ['D5', 'E5', 'D6', 'G8']

    def __init__(self, *args, **kwargs):
        super(Command, self).__init__(*args, **kwargs)
        self.file, self.sheet = self.load_file_and_sheet()

    def handle(self, *args, **options):
        """Open template's file and start proceeding it."""
        self.set_collapse_controls()
        self.fill_header()
        self.write_catalog()
        self.hide_formulas()
        self.set_styles()
        base_dir = settings.ASSETS_DIR
        self.file.save(os.path.join(base_dir, self.NAME))

    def set_styles(self):
        for cell in self.BAD_STYLED_CELLS:
            self.sheet[cell].border = self.THIN_BORDER

    def set_collapse_controls(self):
        """
        Place collapse buttons above rows.

        Collapse controls looks like this: http://prntscr.com/clf9xh. # Ignore InvalidLinkBear
        Doc link: https://goo.gl/nR5pLO
        """
        self.sheet.sheet_properties.outlinePr.summaryBelow = False

    def increase_row(self):
        self.CURRENT_ROW = str(int(self.CURRENT_ROW) + 1)
        return self.CURRENT_ROW

    def get_row(self, row_number):
        return self.sheet.row_dimensions[int(row_number)]

    def load_file_and_sheet(self):
        """
        Load template file into openpyxl.

        Return tuple with opened openpyxl file's object and active price sheet.
        """
        file = openpyxl.load_workbook(os.path.join(
            settings.BASE_DIR, self.TEMPLATE))
        return file, file.get_sheet_by_name('Прайслист')

    def fill_header(self):
        """Fill header of a sheet with date and title."""
        date_cell = 'C5'
        self.sheet.title = self.SHEET_TITLE
        self.sheet[date_cell] = datetime.date.strftime(
            datetime.date.today(), '%d.%m.%Y')

    def hide_formulas(self):
        """Hide formulas for calculating totals."""
        self.sheet.column_dimensions.group('H', 'K', hidden=True, outline_level=0)

    def write_catalog(self):
        """Write categories and products to sheet."""
        categories = Category.objects.active().order_by('name').filter(children=None)
        for category in categories.iterator():
            self.write_category_with_products(category)

    def write_category_with_products(self, category):
        """Write category line and beside that - all of products in this category."""
        def hide_row(row):
            row.hidden = True
            row.outlineLevel = 1

        def collapse_row(row):
            row.collapsed = True

        def write_product_rows():
            """Write products lines."""
            sheet = self.sheet
            products = Product.objects.filter(category=category, page__is_active=True)
            for product in products.iterator():
                product_start = 'A' + self.CURRENT_ROW
                sheet[product_start] = product.name
                sheet[product_start].font = Font(color=colors.BLUE)
                sheet[product_start].hyperlink = settings.BASE_URL + product.url
                sheet[product_start].border = self.THIN_BORDER
                prices = [
                    product.price,
                    product.wholesale_small,
                    product.wholesale_medium,
                    product.wholesale_large,
                ]
                for price, total in zip('CDEF', 'HIJK'):
                    sheet[price + self.CURRENT_ROW] = prices.pop(0)
                    sheet[total + self.CURRENT_ROW] = (
                        '={0}{1}*G{1}'.format(price, self.CURRENT_ROW)
                    )

                    sheet[price + self.CURRENT_ROW].border = self.THIN_BORDER

                sheet['G' + self.CURRENT_ROW].fill = self.BUY_FILL
                sheet['G' + self.CURRENT_ROW].border = self.THIN_BORDER

                hide_row(self.get_row(self.CURRENT_ROW))
                self.increase_row()

        def write_category_row():
            """Merge category line into one cell and write to it."""
            sheet = self.sheet
            collapse_row(self.get_row(self.CURRENT_ROW))

            category_start = 'A' + self.CURRENT_ROW
            category_line = '{}:{}'.format(
                category_start, 'G' + self.CURRENT_ROW)
            sheet.merge_cells(category_line)
            sheet[category_start] = category.name
            sheet[category_start].fill = self.CATEGORY_FILL

            self.increase_row()

        write_category_row()
        write_product_rows()
